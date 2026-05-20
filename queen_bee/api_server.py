"""API服务器 — FastAPI微服务"""
import json, time, asyncio
from pathlib import Path
from .agent import create_agent, Agent
from .config import get as get_config
from .models import ChatRequest, ChatResponse, StatusResponse, SwitchModelRequest, ModelConfigRequest, DeleteModelRequest

try:
    import psutil; HAS_PSUTIL = True
except ImportError:
    HAS_PSUTIL = False

try:
    from fastapi import FastAPI, HTTPException, Request, UploadFile, File as FastAPIFile
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.staticfiles import StaticFiles
    from fastapi.responses import FileResponse
    import uvicorn
except ImportError:
    FastAPI = None

app = FastAPI(title="Eva Agent", version="0.11.5") if FastAPI else None
_agent: Agent = None
ASSETS_DIR = Path.home() / "eva_assets"

def get_agent() -> Agent:
    global _agent
    if _agent is None:
        _agent = create_agent()
    return _agent

# ── 路由 ──

if FastAPI:
    app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

    @app.get("/api/health")
    async def health_check():
        return {"status": "ok", "version": "0.11.5"}

    @app.post("/chat", response_model=ChatResponse)
    async def chat_endpoint(req: ChatRequest):
        try:
            agent = get_agent()
            avatar_name = nova_router.route(req.message)
            if avatar_name:
                av = Agent.with_avatar(avatar_name)
                if av:
                    reply = f"【{get_avatar(avatar_name)['name']}】{av.run(req.message)}"
                else:
                    reply = agent.run(req.message)
            else:
                reply = agent.run(req.message)
        except Exception as e:
            return ChatResponse(reply=f"错误: {e}", timestamp=time.time())
        return ChatResponse(reply=reply, timestamp=time.time())

    @app.get("/status", response_model=StatusResponse)
    async def status_endpoint():
        config = get_config()
        if HAS_PSUTIL:
            mem = psutil.virtual_memory()
            cpu = psutil.cpu_percent(interval=0.1)
            mem_pct = mem.percent
            mem_used_mb = mem.used / 1048576
            mem_total_mb = mem.total / 1048576
        else:
            try:
                with open("/proc/meminfo") as f:
                    raw = f.read()
                lines = raw.splitlines()
                total = int([l for l in lines if "MemTotal" in l][0].split()[1])
                avail = int([l for l in lines if "MemAvailable" in l][0].split()[1])
                mem_pct = round(100 * (total - avail) / total, 1)
                mem_total_mb = total / 1024
                mem_used_mb = (total - avail) / 1024
            except Exception:
                mem_pct = mem_used_mb = mem_total_mb = 0
            try:
                with open("/proc/stat") as f:
                    lines = f.readlines()
                prev = [int(x) for x in [l for l in lines if l.startswith("cpu ")][0].split()[1:]]
                await asyncio.sleep(0.2)
                with open("/proc/stat") as f:
                    lines = f.readlines()
                cur = [int(x) for x in [l for l in lines if l.startswith("cpu ")][0].split()[1:]]
                delta_total = sum(cur) - sum(prev)
                delta_idle = (cur[3] + cur[4]) - (prev[3] + prev[4])
                cpu = round(100 * (1 - delta_idle / max(delta_total, 1)), 1)
            except Exception:
                cpu = 0.0
        return StatusResponse(cpu_percent=cpu, memory_percent=mem_pct, memory_used_mb=mem_used_mb, memory_total_mb=mem_total_mb, agent_ready=_agent is not None, model=f"{config['model']['provider']}/{config['model']['name']}", version="0.11.5")

    @app.get("/api/i18n/zh")
    async def i18n_zh():
        return json.loads((Path(__file__).parent / "locales" / "zh.json").read_text(encoding="utf-8"))

    @app.get("/api/i18n/en")
    async def i18n_en():
        return json.loads((Path(__file__).parent / "locales" / "en.json").read_text(encoding="utf-8"))

    @app.post("/reset")
    async def reset_endpoint():
        get_agent().reset(); return {"status": "ok"}

    @app.get("/models")
    async def models_endpoint():
        from .config import list_models; return list_models()

    @app.post("/model/switch")
    async def switch_model_endpoint(req: SwitchModelRequest):
        from .config import set_model
        try:
            set_model(req.provider, req.name)
            global _agent; _agent = None
            return {"status": "ok", "provider": req.provider, "name": req.name}
        except ValueError as e:
            raise HTTPException(400, str(e))

    @app.get("/config")
    async def config_endpoint():
        cfg = get_config()
        model = dict(cfg.get("model", {}))
        if model.get("api_key"):
            model["api_key"] = model["api_key"][:8] + "****"
        api_keys = {k: v[:8]+"****" for k, v in cfg.get("api_keys", {}).items() if v}
        return {"current": {"provider": model.get("provider",""), "name": model.get("name","")}, "current_model": model, "models_available": cfg.get("models_available", {}), "api_keys": api_keys, "eva_name": cfg.get("eva_name", ""), "data_dir": cfg.get("data_dir", "~/eva-data"), "lab_joined": cfg.get("lab_joined", false)}

    @app.post("/config/model")
    async def config_model_endpoint(req: ModelConfigRequest):
        from .config import save_config, list_models
        cfg = get_config()
        avail = cfg.setdefault("models_available", {})
        avail.setdefault(req.provider, {})[req.name] = {"base_url": req.base_url}
        if req.api_key:
            cfg.setdefault("api_keys", {})[req.provider] = req.api_key
            cfg.setdefault("model", {})["api_key"] = req.api_key
        if not cfg["model"].get("provider"):
            cfg["model"]["provider"] = req.provider; cfg["model"]["name"] = req.name
        save_config()
        return {"status": "ok", "models": list_models()}

    @app.post("/config/model/delete")
    async def config_model_delete_endpoint(req: DeleteModelRequest):
        from .config import save_config
        cfg = get_config()
        avail = cfg.get("models_available", {})
        p, n = req.provider, req.name
        if p not in avail or n not in avail[p]:
            raise HTTPException(404, "模型不存在")
        del avail[p][n]
        if not avail[p]:
            del avail[p]
        if cfg["model"].get("provider") == p and cfg["model"].get("name") == n:
            cfg["model"]["provider"] = ""; cfg["model"]["name"] = ""
        save_config()
        return {"status": "ok", "deleted": f"{p}/{n}"}

    @app.get("/agents")
    async def list_agents_endpoint():
        from .delegate import get_manager; return {"agents": get_manager().list_agents()}

    @app.get("/mastery")
    async def mastery_endpoint():
        from .mastery import score; return score()

    @app.get("/api/memories")
    async def memories_endpoint():
        from .palace import get_palace
        return {"memories": get_palace().recall("")[:10]}

    @app.get("/api/tools")
    async def tools_endpoint():
        from .tool_registry import get_schemas
        tools = get_schemas()
        return {"tools": [{"name": t["function"]["name"], "desc": t["function"]["description"][:80]} for t in tools]}

    @app.get("/api/files")
    async def files_endpoint(dir: str = ""):
        base = Path(dir.replace("~", str(Path.home()))).expanduser() if dir else Path.home() / "eva-data"
        result = []
        if base.exists():
            for f in sorted(base.rglob("*"))[:60]:
                if f.is_file() and not f.name.startswith("."):
                    rel = str(f.relative_to(base))
                    result.append({"name": f.name, "path": rel, "size": f.stat().st_size})
        return {"files": result, "root": str(base)}

    @app.post("/api/files/mkdir")
    async def files_mkdir_endpoint(dir: str = "", name: str = ""):
        base = Path(dir.replace("~", str(Path.home()))).expanduser() if dir else Path.home() / "eva-data"
        (base / name).mkdir(parents=True, exist_ok=True)
        return {"status": "ok"}

    @app.get("/api/dirs")
    async def dirs_endpoint(path: str = ""):
        base = Path(path.replace("~", str(Path.home()))).expanduser() if path else Path.home()
        dirs = []
        if base.exists() and base.is_dir():
            for d in sorted(base.iterdir()):
                if d.is_dir() and not d.name.startswith("."):
                    dirs.append({"name": d.name, "path": str(d)})
        return {"dirs": dirs, "current": str(base), "parent": str(base.parent) if base.parent != base else ""}

    @app.get("/api/files/read")
    async def files_read_endpoint(path: str = "", dir: str = ""):
        base = Path(dir.replace("~", str(Path.home()))).expanduser() if dir else Path.home() / "eva-data"
        fp = base / path
        if not fp.exists() or not fp.is_file():
            return {"content": "文件不存在"}
        ext = fp.suffix.lower()
        try:
            if ext == ".pdf":
                from pdfminer.high_level import extract_text
                text = extract_text(fp)[:5000]
            elif ext in (".docx", ".doc"):
                from docx import Document
                doc = Document(fp)
                text = "\n".join(p.text for p in doc.paragraphs)[:5000]
            elif ext in (".xlsx", ".xls"):
                from openpyxl import load_workbook
                wb = load_workbook(fp, read_only=True, data_only=True)
                lines = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        lines.append("\t".join(str(c or "") for c in row))
                text = "\n".join(lines)[:5000]
            else:
                text = fp.read_text(errors="replace")[:5000]
            return {"content": text}
        except ImportError:
            return {"content": f"读取 {ext} 文件需要安装依赖：pip install pdfminer.six python-docx openpyxl"}
        except Exception as e:
            return {"content": f"读取失败: {e}"}

    @app.post("/api/v1/lab/register")
    async def lab_register_endpoint(req: Request):
        import httpx
        data = await req.json()
        try:
            async with httpx.AsyncClient(timeout=10) as c:
                r = await c.post("http://101.42.23.215:19197/api/v1/lab/register", json={"name": data.get("name",""), "source": "eva-agent"})
                return r.json()
        except Exception:
            return {"member_id": "QL-0000", "status": "offline"}

    @app.post("/config/set")
    async def config_set_endpoint(req: Request):
        from .config import save_config
        data = await req.json()
        cfg = get_config()
        for k, v in data.items():
            cfg[k] = v
        save_config()
        return {"status": "ok"}

    @app.get("/api/sessions")
    async def sessions_endpoint():
        from .mastery import _load as load_mastery
        sessions_dir = Path.home() / ".eva" / "sessions"
        sessions_dir.mkdir(parents=True, exist_ok=True)
        sessions = []
        for f in sorted(sessions_dir.glob("*.json"), reverse=True)[:20]:
            try:
                s = json.loads(f.read_text())
                sessions.append({"id": f.stem, "name": s.get("name", "未命名"), "updated": s.get("updated", ""), "count": s.get("count", 0)})
            except: pass
        return {"sessions": sessions}

    @app.post("/api/sessions")
    async def create_session_endpoint(req: Request):
        data = await req.json()
        sid = str(int(time.time()))
        session = {"name": data.get("name", "新对话"), "id": sid, "created": time.time(), "updated": time.time(), "count": 0}
        sessions_dir = Path.home() / ".eva" / "sessions"
        sessions_dir.mkdir(parents=True, exist_ok=True)
        (sessions_dir / f"{sid}.json").write_text(json.dumps(session, ensure_ascii=False))
        return {"id": sid, "name": session["name"]}

    @app.get("/api/version/check")
    async def version_check_endpoint():
        try:
            import httpx
            r = httpx.get("https://api.github.com/repos/liunian-qianye/eva-agent/releases/latest",
                         timeout=5, headers={"Accept": "application/json"})
            if r.status_code == 200:
                data = r.json()
                latest = data.get("tag_name", "").lstrip("v")
                current = "0.11.3"
                return {"current": current, "latest": latest, "update_available": latest > current,
                        "url": data.get("html_url", "")}
        except Exception:
            pass
        return {"current": "0.11.5", "latest": "", "update_available": False}

    # 服务 React 前端（放在所有 API 路由之后）
    react_dist = Path(__file__).parent.parent / "desktop-ui" / "dist"
    if react_dist.exists():
        from fastapi.staticfiles import StaticFiles
        from fastapi.responses import FileResponse
        app.mount("/assets", StaticFiles(directory=str(react_dist / "assets")), name="assets")

        @app.get("/{full_path:path}")
        async def serve_react(full_path: str):
            return FileResponse(str(react_dist / "index.html"))

def run_server(host: str = None, port: int = None, reload: bool = False):
    if not FastAPI:
        print("需要安装: pip install fastapi uvicorn psutil"); return
    config = get_config()
    host = host or config["server"]["host"]; port = port or config["server"]["port"]
    get_agent()
    print(f"Eva Agent v0.11.5 启动 — 千叶实验室")
    print(f"桌面端: http://localhost:{port}")
    print(f"API文档: http://localhost:{port}/docs")
    print(f"热重载: {'开启' if reload else '关闭'} — 改代码后自动重启")
    uvicorn.run(app, host=host, port=port, reload=reload)
